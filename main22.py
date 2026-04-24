import PySimpleGUI as sg
from openpyxl import Workbook,load_workbook
import tkinter as tk
import os

template_file = 'Accounting_Template.xlsx'
save_file = 'Accounting01.xlsx'
u_font=('Yu Gothic UI Semibold', 14)
accounting_items = ['交付金','振興費','会議費','雑収入','指導派遣費','通信事務所',' ','その他']


def create_menu_window():
    layout = [
        [sg.Text('メニュー', font=('Yu Gothic UI Semibold', 20))],
        [sg.Button('入力画面へ', size=(20,2))],
        [sg.Button('履歴を見る', size=(20,2))],
        [sg.Button('終了', size=(20,2))]

    ]
    return sg.Window('メニュー', layout, size=(400,300) ,finalize=True, font=u_font)


def create_history_window(data):
    headings = ['日付', '項目', '内容', '収入金額', '支払金額', '差引金額']

    layout = [
        [sg.Table(
            values=data,
            headings=headings,
            key = 'table',
            enable_events=True,
            auto_size_columns=True,
            justification='center',
            num_rows=10
        )],
        [sg.Button('編集'), sg.Button('削除'), sg.Button('戻る')]

    ]
    return sg.Window('履歴', layout, size=(800,600), finalize=True, font=u_font)


def create_input_window():


    
    layout = [
        [sg.Text('日付：', size=(12,1),justification='right'),
        sg.Input(key='text_date', size=(20,1)),
        sg.Push(),
        sg.CalendarButton('日付選択', format='%m-%d',key='button_calendar', target='text_date', size=(20,1))],

        [sg.Text('項目：', size=(12,1),justification='right'), sg.Listbox(values=accounting_items,key='list_item', 
                                    enable_events=True, size=(40,5))],

        [sg.Text('その他入力：', size=(12,1),justification='right'),sg.Input(key='other', disabled=True, size=(30,1))],

        [sg.Text('内容：', size=(12,1),justification='right'), sg.InputText(key='input_content', size=(30,1))],

        [sg.Text('収入金額：', size=(12,1),justification='right'), sg.InputText(key='input_income', enable_events=True, size=(20,1))],

        [sg.Text('支払金額：', size=(12,1),justification='right'), sg.InputText(key='input_expense', enable_events=True, size=(20,1))],

        [sg.Text('差引金額：', size=(12,1),justification='right'), sg.InputText(key='input_net',disabled=True, size=(20,1))],

        [sg.Push(),
        sg.Button('保存して次へ', size=(15,1)), 
        sg.Button('戻る', size=(15,1))]
        ]
    return sg.Window('帳簿入力',layout,size=(800,600), finalize=True, font=u_font, resizable=True)
    finalize=True 


def is_number(s):
    try:
        int(s)
        return True
    except ValueError:
        return False
    


#Excelを準備    
wb = load_workbook(template_file)
ws = wb.active
row =  4

#事前変数
edit_index = None #編集モードのインデックス
history_data = []
current_window = create_menu_window()

sg.theme('TealMono')


while True:
    event, values = current_window.read()
    
    if event in (sg.WIN_CLOSED, '終了'):
        break
    #--------
    #画面遷移
    #--------

    if event == '入力画面へ':
        current_window.close()
        current_window = create_input_window() #最初に作った入力画面へ

    if event == '履歴を見る':
        current_window.close()
        current_window = create_history_window(history_data) #履歴画面へ    

    elif event == '戻る':
        current_window.close()
        current_window = create_menu_window() #戻る

    #--------
    #履歴編集
    #--------

    elif event == '編集':
        selected = values['table']
        if not selected:
            sg.popup('選択してください' \
            '')
            continue
        
        edit_index = selected[0]
        record = history_data[edit_index]

        current_window.close()
        current_window = create_input_window()

        #入力画面に履歴の値をセット
        current_window['text_date'].update(value=record[0])
        current_window['input_content'].update(value=record[2])
        current_window['input_income'].update(value=(record[3]))
        current_window['input_expense'].update(value=(record[4]))
        current_window['input_net'].update(value=(record[5]))

        if record[1] in accounting_items:
            current_window['list_item'].update(set_to_index=[accounting_items.index(record[1])])
        else:
            current_window['list_item'].update(set_to_index=[accounting_items.index('その他')])
            current_window['other'].update(value=record[1], disabled=False)

    #--------
    #入力画面の処理
    #--------

    #その他が選ばれたら入力欄を有効化
    if event == 'list_item':
        if values['list_item'] and values['list_item'][0] == 'その他':
            current_window['other'].update(disabled=False)
        else:
            current_window['other'].update(disabled=True)
            current_window['other'].update(value='') #リセット

    #差引金額を計算して表示       
    if event in ('input_income', 'input_expense'):
            income = values['input_income']
            expense = values['input_expense']
            if income.isdigit() and expense.isdigit():
                net = int(income) - int(expense)
                current_window['input_net'].update(value=str(net))
            else:
                current_window['input_net'].update('')

    #保存して次へ
    if event == '保存して次へ':
        
        #キーの内容
        entry_content = values['input_content']
        entry_calander = values['text_date']
        income = values['input_income']
        expense = values['input_expense']

        #入力内容のバリデーション
        if not entry_calander:
            sg.popup('日付を選択してください')
            continue

        if not values['list_item']:
            sg.popup('項目を選択してください')
            continue

        if not values['input_content']:
            sg.popup('内容を入力してください')
            continue

        if values['list_item']:
            if values['list_item'][0] == 'その他':
                if not values['other']:
                    sg.popup('その他の項目を入力してください')
                    continue
                choose_item = values['other'] #入力された値
            else:
                choose_item = values['list_item'][0] # 選択された値

        if not is_number(income) or not is_number(expense):
            sg.popup('金額は数字で入力してください')
            continue

        #保存分岐
        if edit_index is not None:
            history_data[edit_index] = [
                entry_calander,
                choose_item,
                entry_content,
                int(income),
                int(expense),
                0
            ]
            sg.popup('更新しました')
            edit_index = None #編集モード終了
        else:
            history_data.append([
                entry_calander,
                choose_item,
                entry_content,
                int(income),
                int(expense),
                0
            ])
            sg.popup('保存しました')

        #Excelに保存
        balance = 0
        start_row = 4
        for i, data in enumerate(history_data):
            ws.cell(row=start_row + i, column=1, value=data[0])
            ws.cell(row=start_row + i, column=2, value=data[1])
            ws.cell(row=start_row + i, column=3, value=data[2])
            ws.cell(row=start_row + i, column=4, value=data[3])
            ws.cell(row=start_row + i, column=5, value=data[4])
            ws.cell(row=start_row + i, column=6, value=data[5])

        wb.save(save_file)
        sg.popup(f"{start_row + len(history_data) - 1}行目に 保存しました")

        current_window['text_date'].update(value='')
        current_window['list_item'].update(set_to_index=[])
        current_window['other'].update(value='', disabled=True)
        current_window['input_content'].update(value='')
        current_window['input_income'].update(value='')
        current_window['input_expense'].update(value='')
        current_window['input_net'].update(value='')
    
    




current_window.close()