from openpyxl import load_workbook
from db.database import history_data
import datetime

template_file = './templates/Accounting_Template.xlsx'
date_created = datetime.datetime.now().strftime('%Y%m%d_%H')
save_file = f'Accounting+{date_created}.xlsx'

def save_to_excel(history_data):
    wb = load_workbook(template_file)
    ws = wb.active

    start_row = 4

    for i, data in enumerate(history_data):
        ws.cell(row=start_row + i, column=1, value=data[0])
        ws.cell(row=start_row + i, column=2, value=data[1])
        ws.cell(row=start_row + i, column=3, value=data[2])
        ws.cell(row=start_row + i, column=4, value=data[3])
        ws.cell(row=start_row + i, column=5, value=data[4])
        ws.cell(row=start_row + i, column=6, value=data[5])

    wb.save(save_file)